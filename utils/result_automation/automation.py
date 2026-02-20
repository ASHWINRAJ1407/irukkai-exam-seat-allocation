import os
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PASS_MARKS = {
    "IA1": 30,
    "IA2": 30,
    "MODEL": 50
}

MAX_MARKS = {
    "IA1": 60,
    "IA2": 60,
    "MODEL": 100
}


def extract_header_details(pdf):

    text = pdf.pages[0].extract_text()
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    details = {
        "department": "",
        "academic_year": "",
        "subject_code": "",
        "subject_name": ""
    }

    for line in lines:
        upper = line.upper()

        if "DEPARTMENT" in upper:
            details["department"] = line

        if "ACADEMIC YEAR" in upper:
            details["academic_year"] = line

        if "COURSE CODE" in upper and "&" in line:
            parts = line.split(":")
            if len(parts) > 1:
                content = parts[1].split("TOTAL")[0].strip()
                if "&" in content:
                    code, name = content.split("&", 1)
                    details["subject_code"] = code.strip()
                    details["subject_name"] = name.strip()

    return details


def extract_students(pdf):

    students = []

    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue

        lines = [l.strip() for l in text.split("\n") if l.strip()]

        for line in lines:
            parts = line.split()

            if len(parts) < 4:
                continue

            if not parts[0].replace(".", "").isdigit():
                continue

            if not parts[1].isdigit():
                continue

            reg_no = parts[1]
            mark = parts[-1]

            if not (mark.isdigit() or mark == "AB"):
                continue

            name = " ".join(parts[2:-1])

            students.append({
                "reg_no": reg_no,
                "name": name.strip(),
                "mark": mark
            })

    return students


def generate_report_flask(exam_type, college_name, input_folder, output_folder):

    pass_mark = PASS_MARKS[exam_type]
    max_mark = MAX_MARKS[exam_type]

    pdf_files = [f for f in os.listdir(input_folder) if f.endswith(".pdf")]
    if not pdf_files:
        return None

    master_data = {}
    subjects = []
    department = ""
    academic_year = ""

    for file in pdf_files:

        with pdfplumber.open(os.path.join(input_folder, file)) as pdf:

            header = extract_header_details(pdf)
            students = extract_students(pdf)

            if not students:
                continue

            subjects.append((header["subject_code"], header["subject_name"]))
            department = header["department"]
            academic_year = header["academic_year"]

            for student in students:
                reg = student["reg_no"]

                if reg not in master_data:
                    master_data[reg] = {
                        "Reg No": reg,
                        "Name": student["name"]
                    }

                master_data[reg][header["subject_code"]] = student["mark"]

    if not master_data:
        return None

    df = pd.DataFrame(master_data.values()).sort_values("Reg No")

    for sub_code, _ in subjects:
        if sub_code not in df.columns:
            df[sub_code] = "AB"

    df = df.fillna("AB")

    df["Total"] = df.apply(
        lambda r: sum(int(r[sub]) for sub, _ in subjects if r[sub] != "AB"),
        axis=1
    )

    df["Mark %"] = (df["Total"] / (len(subjects) * max_mark)) * 100

    def result_calc(row):
        p = f = a = 0
        for sub, _ in subjects:
            m = row[sub]
            if m == "AB":
                a += 1
            elif int(m) >= pass_mark:
                p += 1
            else:
                f += 1
        return pd.Series([p, f, a])

    df[["Pass", "Fail", "AB"]] = df.apply(result_calc, axis=1)

    df["Rank"] = ""
    eligible = df[(df["Fail"] == 0) & (df["AB"] == 0)]

    if not eligible.empty:
        ranks = eligible["Total"].rank(method="min", ascending=False)
        for idx in eligible.index:
            df.at[idx, "Rank"] = str(int(ranks[idx]))

    wb = Workbook()
    ws = wb.active
    ws.title = exam_type

    fail_fill = PatternFill(start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid")

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cols = 3 + len(subjects) + 6
    row_no = 1

    # HEADER
    header_lines = [
        college_name,
        department,
        f"{exam_type} - CONSOLIDATED MARK REPORT",
        academic_year
    ]

    for text in header_lines:
        ws.merge_cells(start_row=row_no, start_column=1,
                       end_row=row_no, end_column=total_cols)
        cell = ws.cell(row=row_no, column=1)
        cell.value = text
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        row_no += 1

    row_no += 2

    # COLUMN HEADERS
    headers = ["S.No", "Reg No", "Name"]
    headers += [code for code, _ in subjects]
    headers += ["Total", "Mark %", "Pass", "Fail", "AB", "Rank"]

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_no, column=col_index)
        cell.value = header
        cell.font = Font(bold=True)

    row_no += 1

    col_index = 4
    for _, name in subjects:
        ws.cell(row=row_no, column=col_index).value = name
        col_index += 1

    row_no += 1

    sno = 1
    for _, row in df.iterrows():

        ws.cell(row=row_no, column=1).value = sno
        ws.cell(row=row_no, column=2).value = row["Reg No"]
        name_cell = ws.cell(row=row_no, column=3)
        name_cell.value = row["Name"]

        col = 4
        failed = row["Fail"] > 0

        for sub_code, _ in subjects:
            val = row[sub_code]
            cell = ws.cell(row=row_no, column=col)
            cell.value = val

            if val != "AB" and int(val) < pass_mark:
                cell.fill = fail_fill

            col += 1

        ws.cell(row=row_no, column=col).value = row["Total"]
        ws.cell(row=row_no, column=col + 1).value = round(row["Mark %"], 2)
        ws.cell(row=row_no, column=col + 2).value = row["Pass"]
        ws.cell(row=row_no, column=col + 3).value = row["Fail"]
        ws.cell(row=row_no, column=col + 4).value = row["AB"]
        ws.cell(row=row_no, column=col + 5).value = row["Rank"]

        if failed:
            name_cell.fill = fail_fill

        sno += 1
        row_no += 1

    # SUMMARY
    row_no += 2

    summary_labels = [
        "Total No. of Students",
        "Total No. of Appeared",
        "Total No. of Absent",
        "Total No. of Passed",
        "Total No. of Failed",
        "Appeared Pass (%)"
    ]

    for label in summary_labels:

        ws.cell(row=row_no, column=1).value = label
        ws.cell(row=row_no, column=1).font = Font(bold=True)

        col = 4

        for sub_code, _ in subjects:

            column_data = df[sub_code]
            numeric_marks = pd.to_numeric(column_data, errors="coerce")

            total_students = len(df)
            appeared = numeric_marks.notna().sum()
            absent = column_data.eq("AB").sum()
            passed = (numeric_marks >= pass_mark).sum()
            failed = appeared - passed

            if label == "Total No. of Students":
                value = total_students
            elif label == "Total No. of Appeared":
                value = appeared
            elif label == "Total No. of Absent":
                value = absent
            elif label == "Total No. of Passed":
                value = passed
            elif label == "Total No. of Failed":
                value = failed
            else:
                value = round((passed / appeared) * 100, 2) if appeared > 0 else 0

            ws.cell(row=row_no, column=col).value = value
            col += 1

        row_no += 1

    # FOOTER
    row_no += 3
    ws.cell(row=row_no, column=2).value = "Staff Signature"

    row_no += 4
    ws.cell(row=row_no, column=2).value = "BATCH COORDINATOR"
    ws.cell(row=row_no, column=total_cols - 1).value = "HoD"

    # FORCE FIXED COLUMN WIDTHS (After all formatting)

    ws.column_dimensions["A"].width = 6      # S.No
    ws.column_dimensions["B"].width = 15     # Reg No
    ws.column_dimensions["C"].width = 25     # Name

    col_index = 4
    for _ in subjects:
        ws.column_dimensions[get_column_letter(col_index)].width = 14
        col_index += 1

    ws.column_dimensions[get_column_letter(col_index)].width = 12   # Total
    ws.column_dimensions[get_column_letter(col_index+1)].width = 12 # %
    ws.column_dimensions[get_column_letter(col_index+2)].width = 8  # Pass
    ws.column_dimensions[get_column_letter(col_index+3)].width = 8  # Fail
    ws.column_dimensions[get_column_letter(col_index+4)].width = 8  # AB
    ws.column_dimensions[get_column_letter(col_index+5)].width = 8  # Rank


    # =========================
    # BORDERS
    # =========================
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # =========================
    # FREEZE HEADER (Better Alignment When Scrolling)
    # =========================
    ws.freeze_panes = "A8"   # Adjust if header rows change

    # =========================
    # FINAL PROFESSIONAL A3 PRINT SETTINGS
    # =========================

    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Fit all columns in 1 page width
    ws.page_setup.fitToWidth = 1

    # Allow multiple pages vertically (front & back)
    ws.page_setup.fitToHeight = False

    # Remove manual scaling
    ws.page_setup.scale = None

    # Center horizontally
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Professional margins
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Define print area
    last_column = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.print_area = f"A1:{last_column}{last_row}"

    # =========================
    # SAVE FILE
    # =========================

    output_path = os.path.join(output_folder,
                               f"{exam_type}_Automation_Report.xlsx")

    wb.save(output_path)

    return output_path
